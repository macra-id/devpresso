import sys
import csv
from PyQt5.QtWidgets import (
    QApplication, QWidget, QMainWindow, QPushButton, QVBoxLayout, QTableWidget,
    QTableWidgetItem, QComboBox, QMessageBox, QInputDialog, QFileDialog, QScrollArea, QHBoxLayout, QDateEdit, QLabel
)
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from PyQt5.QtCore import (QDate, Qt)
from PyQt5.QtGui import QPainter, QFont
from PyQt5.QtPrintSupport import QPrinter, QPrintDialog
import pymysql

# Database connection settings
DB_CONFIG = {
    'host': '127.0.0.1',
    'user': 'root',
    'password': '',
    'database': 'devpresso_db',
    'port': 3306
}

def create_database():
    connection = None
    try:
        print("Connecting to MySQL server...")
        connection = pymysql.connect(
            host=DB_CONFIG['host'],
            user=DB_CONFIG['user'],
            password=DB_CONFIG['password'],
            port=DB_CONFIG['port']
        )
        print("Connected to MySQL server")
        
        cursor = connection.cursor()
        
        print(f"Creating database {DB_CONFIG['database']}...")
        cursor.execute(f"CREATE DATABASE IF NOT EXISTS {DB_CONFIG['database']}")
        print("Database created or already exists")
        
        cursor.execute(f"USE {DB_CONFIG['database']}")
        print("Using database")

        print("Creating transactions table...")
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS transactions (
                id INT AUTO_INCREMENT PRIMARY KEY,
                customer_name VARCHAR(255),
                drink_type VARCHAR(255),
                variant VARCHAR(255),
                quantity INT,
                total_price DECIMAL(10,2),
                date DATE,
                paid BOOLEAN DEFAULT 0,
                payment_method VARCHAR(50) DEFAULT '-'
            )
        """)
        print("Transactions table created")
        
        print("Creating drinks table...")
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS drinks (
                id INT AUTO_INCREMENT PRIMARY KEY,
                drink_type VARCHAR(255),
                variant VARCHAR(255),
                price DECIMAL(10,2)
            )
        """)
        print("Drinks table created")
        
        connection.commit()
        print("Database setup completed successfully")

    except pymysql.MySQLError as err:
        print(f"MySQL Error: {err}")
    finally:
        if connection and connection.open:
            connection.close()
            print("Database connection closed")


class MainMenu(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Devpresso")
        self.setGeometry(100, 100, 400, 200)
        self.setMinimumSize(300, 150)
        self.init_ui()

    def init_ui(self):
        layout = QVBoxLayout()
        layout.setSpacing(20)
        layout.setContentsMargins(30, 30, 30, 30)
        self.create_buttons(layout)
        central_widget = QWidget()
        central_widget.setLayout(layout)
        self.setCentralWidget(central_widget)
        self.apply_styles()

    def create_buttons(self, layout):
        transactions_btn = QPushButton("Transactions", self)
        transactions_btn.clicked.connect(self.open_transactions_window)
        layout.addWidget(transactions_btn)

        drink_menu_btn = QPushButton("Drinks Menu", self)
        drink_menu_btn.clicked.connect(self.open_drink_menu_window)
        layout.addWidget(drink_menu_btn)

    def open_transactions_window(self):
        self.transactions_window = TransactionsWindow()
        self.transactions_window.show()

    def open_drink_menu_window(self):
        self.drink_menu_window = DrinkMenuWindow()
        self.drink_menu_window.show()

    def apply_styles(self):
        self.setStyleSheet("""
            QPushButton {
                font-size: 16px;
                padding: 10px;
                background-color: #007BFF;
                color: white;
                border-radius: 5px;
            }
            QPushButton:hover {
                background-color: #0056b3;
            }
            QMainWindow {
                background-color: #f0f0f0;
            }
        """)

class TransactionsWindow(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Transactions")
        self.setGeometry(150, 150, 800, 500)
        self.setMinimumSize(1000, 400)
        self.init_ui()

    def init_ui(self):
        layout = QVBoxLayout()
        layout.setSpacing(15)
        layout.setContentsMargins(20, 20, 20, 20)

        # Date filter layout
        filter_layout = QHBoxLayout()
        self.setup_date_filters(filter_layout)
        layout.addLayout(filter_layout)

        # Transactions table
        self.setup_transactions_table()
        layout.addWidget(self.transactions_table)

        # Buttons
        btn_layout = self.setup_buttons()
        layout.addLayout(btn_layout)

        self.setLayout(layout)
        self.load_transactions()

    def setup_date_filters(self, filter_layout):
        self.start_date_edit = QDateEdit()
        self.start_date_edit.setDate(QDate.currentDate())
        self.start_date_edit.setCalendarPopup(True)
        filter_layout.addWidget(QLabel("Start Date:"))
        filter_layout.addWidget(self.start_date_edit)

        self.end_date_edit = QDateEdit()
        self.end_date_edit.setDate(QDate.currentDate())
        self.end_date_edit.setCalendarPopup(True)
        filter_layout.addWidget(QLabel("End Date:"))
        filter_layout.addWidget(self.end_date_edit)

        self.filter_btn = QPushButton("Apply Filter")
        self.filter_btn.clicked.connect(self.load_transactions)
        filter_layout.addWidget(self.filter_btn)

    def setup_transactions_table(self):
        self.transactions_table = QTableWidget(0, 10)
        self.transactions_table.setHorizontalHeaderLabels(
            ["No", "Date", "Customer Name", "Drink Type", "Variant", "Quantity", 
             "Total Price (Rp)", "Paid", "Payment Method", "Print"]
        )
        self.transactions_table.horizontalHeader().setStretchLastSection(True)
        self.transactions_table.verticalHeader().setVisible(False)
        self.transactions_table.setAlternatingRowColors(True)
        self.transactions_table.itemChanged.connect(self.handle_item_changed)

    def setup_buttons(self):
        btn_layout = QHBoxLayout()
        btn_layout.setSpacing(10)

        self.add_btn = QPushButton("Add Transaction")
        self.add_btn.clicked.connect(self.add_transaction)
        btn_layout.addWidget(self.add_btn)

        self.delete_btn = QPushButton("Delete Transaction")
        self.delete_btn.clicked.connect(self.delete_transaction)
        btn_layout.addWidget(self.delete_btn)

        self.download_csv_btn = QPushButton("Download CSV")
        self.download_csv_btn.clicked.connect(self.download_transactions_csv)
        btn_layout.addWidget(self.download_csv_btn)

        self.download_excel_btn = QPushButton("Download Excel")
        self.download_excel_btn.clicked.connect(self.download_transactions_excel)
        btn_layout.addWidget(self.download_excel_btn)

        return btn_layout

    def load_transactions(self):
        try:
            print("Connecting to MySQL for loading transactions...")
            connection = pymysql.connect(**DB_CONFIG)
            print("Connected to MySQL for transactions")

            start_date = self.start_date_edit.date().toString("yyyy-MM-dd")
            end_date = self.end_date_edit.date().toString("yyyy-MM-dd")
            print(f"Filtering transactions from {start_date} to {end_date}")

            cursor = connection.cursor()
            cursor.execute("""
                SELECT id, date, customer_name, drink_type, variant, quantity, 
                    total_price, paid, payment_method 
                FROM transactions
                WHERE date BETWEEN %s AND %s
            """, (start_date, end_date))

            records = cursor.fetchall()
            print(f"Fetched transactions: {records}")

            self.transactions_table.setRowCount(0)
            for row_data in records:
                row_count = self.transactions_table.rowCount()
                self.transactions_table.insertRow(row_count)

                # Populate each cell
                for col, data in enumerate(row_data):
                    if col == 7:  # Paid column (Checkbox)
                        checkbox_item = QTableWidgetItem()
                        checkbox_item.setFlags(Qt.ItemIsUserCheckable | Qt.ItemIsEnabled)
                        checkbox_item.setCheckState(Qt.Checked if data else Qt.Unchecked)
                        self.transactions_table.setItem(row_count, col, checkbox_item)
                    elif col == 8:  # Payment Method column (Dropdown)
                        combo = QComboBox()
                        combo.addItems(["QRIS", "Cash", "-"])
                        combo.setCurrentText(str(data))
                        combo.currentIndexChanged.connect(
                            lambda index, r=row_count, c=col: self.handle_payment_method_change(r, c)
                        )
                        self.transactions_table.setCellWidget(row_count, col, combo)
                    elif col == 6:  # Total Price column (Formatted Price)
                        formatted_price = f"{float(data):,.0f}"
                        self.transactions_table.setItem(row_count, col, QTableWidgetItem(formatted_price))
                    else:
                        self.transactions_table.setItem(row_count, col, QTableWidgetItem(str(data)))

                # Add Print button in the last column
                print_btn = QPushButton("Print")
                print_btn.clicked.connect(lambda _, r=row_count: self.print_transaction(r))
                self.transactions_table.setCellWidget(row_count, 9, print_btn)

            print("Transactions loaded successfully")

        except pymysql.MySQLError as e:
            QMessageBox.critical(self, "Database Error", f"Error loading transactions: {e}")
        finally:
            if connection and connection.open:
                cursor.close()
                connection.close()
                print("Database connection closed for transactions")


    def handle_item_changed(self, item):
        row = item.row()
        column = item.column()
        transaction_id = self.transactions_table.item(row, 0).text()

        try:
            connection = pymysql.connect(**DB_CONFIG)
            cursor = connection.cursor()

            if column == 7:  # Paid column
                new_value = 1 if item.checkState() == Qt.Checked else 0
                cursor.execute(
                    "UPDATE transactions SET paid = %s WHERE id = %s",
                    (new_value, transaction_id)
                )
            elif column == 8:  # Payment Method column
                widget = self.transactions_table.cellWidget(row, column)
                if widget is not None:
                    new_value = widget.currentText()
                else:
                    return


            connection.commit()

        except pymysql.MySQLError as e:
            QMessageBox.critical(self, "Database Error", f"Error updating transaction: {e}")
        finally:
            if connection and connection.open:
                cursor.close()
                connection.close()
                
    def handle_payment_method_change(self, row, column):
        transaction_id = self.transactions_table.item(row, 0).text()
        new_value = self.transactions_table.cellWidget(row, column).currentText()

        try:
            connection = pymysql.connect(**DB_CONFIG)
            cursor = connection.cursor()
            cursor.execute(
                "UPDATE transactions SET payment_method = %s WHERE id = %s",
                (new_value, transaction_id)
            )
            connection.commit()
        except pymysql.MySQLError as e:
            QMessageBox.critical(self, "Database Error", f"Error updating payment method: {e}")
        finally:
            if connection and connection.open:
                cursor.close()
                connection.close()

    def add_transaction(self):
        try:
            connection = pymysql.connect(**DB_CONFIG)
            cursor = connection.cursor()

            # Get customer name
            customer_name, ok0 = QInputDialog.getText(self, "Add Transaction", "Enter Customer Name:")
            if not ok0 or not customer_name.strip():
                return

            # Get drink types
            cursor.execute("SELECT DISTINCT drink_type FROM drinks")
            drink_types = [row[0] for row in cursor.fetchall()]
            
            if not drink_types:
                QMessageBox.warning(self, "No Drinks Available", "Please add drinks to the menu first.")
                return

            drink_type, ok1 = QInputDialog.getItem(
                self, "Add Transaction", "Select Drink Type:", 
                drink_types, editable=False
            )
            if not ok1:
                return

            # Get variants
            cursor.execute(
                "SELECT variant, price FROM drinks WHERE drink_type = %s",
                (drink_type,)
            )
            variants = cursor.fetchall()
            
            if not variants:
                QMessageBox.warning(
                    self, "No Variants Available",
                    f"No variants found for the drink type '{drink_type}'."
                )
                return

            variant_options = [f"{variant} - Rp {price:,.0f}" for variant, price in variants]
            selected_variant, ok2 = QInputDialog.getItem(
                self, "Add Transaction", "Select Variant:",
                variant_options, editable=False
            )
            if not ok2:
                return

            variant, price_str = selected_variant.split(" - ")
            price = float(price_str.replace("Rp ", "").replace(",", ""))

            quantity, ok3 = QInputDialog.getInt(self, "Add Transaction", "Enter Quantity:", min=1)
            if not ok3:
                return

            total_price = price * quantity
            transaction_date = QDate.currentDate().toString("yyyy-MM-dd")

            cursor.execute("""
                INSERT INTO transactions 
                (customer_name, drink_type, variant, quantity, total_price, date)
                VALUES (%s, %s, %s, %s, %s, %s)
            """, (customer_name, drink_type, variant, quantity, total_price, transaction_date))

            connection.commit()
            self.load_transactions()

        except pymysql.MySQLError as e:
            QMessageBox.critical(self, "Database Error", f"Error adding transaction: {e}")
        finally:
            if connection and connection.open:
                cursor.close()
                connection.close()

    def delete_transaction(self):
        selected_row = self.transactions_table.currentRow()
        if selected_row < 0:
            QMessageBox.warning(self, "Delete Transaction", "Please select a transaction to delete.")
            return

        transaction_id = self.transactions_table.item(selected_row, 0).text()

        confirmation = QMessageBox.question(
            self,
            "Confirm Deletion",
            f"Are you sure you want to delete the transaction with ID '{transaction_id}'?",
            QMessageBox.Yes | QMessageBox.No
        )

        if confirmation == QMessageBox.Yes:
            try:
                connection = pymysql.connect(**DB_CONFIG)
                cursor = connection.cursor()
                cursor.execute("DELETE FROM transactions WHERE id = %s", (transaction_id,))
                connection.commit()
                self.load_transactions()
            except pymysql.MySQLError as e:
                QMessageBox.critical(self, "Database Error", f"Error deleting transaction: {e}")
            finally:
                if connection and connection.open:
                    cursor.close()
                    connection.close()
                    
    def print_transaction(self, row):
        transaction_id = self.transactions_table.item(row, 0).text()
        customer_name = self.transactions_table.item(row, 2).text()
        drink_type = self.transactions_table.item(row, 3).text()
        variant = self.transactions_table.item(row, 4).text()
        quantity = self.transactions_table.item(row, 5).text()
        total_price = self.transactions_table.item(row, 6).text()
        paid = self.transactions_table.item(row, 7).checkState() == Qt.Checked
        payment_method = self.transactions_table.cellWidget(row, 8).currentText()

        # Construct the receipt text
        receipt = f"""
        Transaction ID: {transaction_id}
        Customer Name: {customer_name}
        Drink: {drink_type} - {variant}
        Quantity: {quantity}
        Total Price: Rp {total_price}
        Paid: {'Yes' if paid else 'No'}
        Payment Method: {payment_method}
        """

        # Print the receipt using QPrinter
        printer = QPrinter()
        dialog = QPrintDialog(printer, self)
        if dialog.exec_() == QPrintDialog.Accepted:
            painter = QPainter(printer)
            painter.setFont(QFont("Arial", 12))
            painter.drawText(100, 100, receipt)
            painter.end()

                    
    def get_filtered_dates(self):
        start_date = self.start_date_edit.date().toString("yyyy-MM-dd")
        end_date = self.end_date_edit.date().toString("yyyy-MM-dd")
        return start_date, end_date

    def download_transactions_csv(self):
        start_date, end_date = self.get_filtered_dates()

        options = QFileDialog.Options()
        file_name, _ = QFileDialog.getSaveFileName(self, "Save Transactions", "", "CSV Files (*.csv);;All Files (*)", options=options)
        if file_name:
            try:
                connection = pymysql.connect(**DB_CONFIG)
                cursor = connection.cursor()
                cursor.execute("""
                    SELECT date, customer_name, drink_type, variant, quantity, total_price, 
                        CASE WHEN paid = 1 THEN 'True' ELSE 'False' END AS paid, payment_method
                    FROM transactions
                    WHERE date BETWEEN %s AND %s
                """, (start_date, end_date))

                records = cursor.fetchall()

                # Calculating totals
                total_quantity = sum(int(record[4]) for record in records)
                total_amount = sum(float(record[5]) for record in records)
                total_paid_true = sum(1 for record in records if record[6] == 'True')

                # Writing data to CSV file
                with open(file_name, mode="w", newline="") as file:
                    writer = csv.writer(file)
                    writer.writerow(["Date", "Customer Name", "Drink Type", "Variant", "Quantity", "Total Price (Rp)", "Paid", "Payment Method"])
                    for row in records:
                        formatted_row = list(row)
                        formatted_row[5] = int(formatted_row[5])  # Convert total_price to integer
                        writer.writerow(formatted_row)

                    # Add totals row
                    paid_ratio = f"{total_paid_true}/{len(records)}"
                    total_row = ["", "", "", "Total", total_quantity, int(total_amount), paid_ratio, ""]
                    writer.writerow(total_row)

                QMessageBox.information(self, "Success", f"Transactions saved to {file_name}")

            except pymysql.MySQLError as e:
                QMessageBox.critical(self, "Database Error", f"Error fetching transactions: {e}")
            finally:
                if connection and connection.open:
                    cursor.close()
                    connection.close()
                    
    def download_transactions_excel(self):
        start_date, end_date = self.get_filtered_dates()

        options = QFileDialog.Options()
        file_name, _ = QFileDialog.getSaveFileName(self, "Save Transactions", "", "Excel Files (*.xlsx);;All Files (*)", options=options)
        if file_name:
            try:
                connection = pymysql.connect(**DB_CONFIG)
                cursor = connection.cursor()
                cursor.execute("""
                    SELECT date, customer_name, drink_type, variant, quantity, total_price, 
                        CASE WHEN paid = 1 THEN 'True' ELSE 'False' END AS paid, payment_method
                    FROM transactions
                    WHERE date BETWEEN %s AND %s
                """, (start_date, end_date))

                records = cursor.fetchall()

                # Calculating totals
                total_quantity = sum(int(record[4]) for record in records)
                total_amount = sum(float(record[5]) for record in records)
                total_paid_true = sum(1 for record in records if record[6] == 'True')

                # Writing data to Excel file
                workbook = Workbook()
                sheet = workbook.active
                sheet.title = "Transactions"

                # Add header row
                headers = ["Date", "Customer Name", "Drink Type", "Variant", "Quantity", "Total Price (Rp)", "Paid", "Payment Method"]
                sheet.append(headers)

                # Style the header row
                header_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF")
                header_alignment = Alignment(horizontal="center")

                for col_num, header in enumerate(headers, start=1):
                    cell = sheet.cell(row=1, column=col_num)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = header_alignment

                # Data rows
                for row in records:
                    formatted_row = list(row)
                    formatted_row[5] = int(formatted_row[5])  # Convert total_price to integer
                    sheet.append(formatted_row)

                # Total row at the bottom
                paid_ratio = f"{total_paid_true}/{len(records)}"
                total_row = ["", "", "", "Total", total_quantity, int(total_amount), paid_ratio, ""]
                sheet.append(total_row)

                # Apply formatting for thousands separator
                for row_idx in range(2, len(records) + 3):  # Rows with data and total
                    sheet.cell(row=row_idx, column=6).number_format = '#,##0'  # Format Total Price (Rp)
                sheet.cell(row=len(records) + 3, column=5).number_format = '#,##0'  # Format Total Quantity
                sheet.cell(row=len(records) + 3, column=6).number_format = '#,##0'  # Format Total Amount

                workbook.save(file_name)
                QMessageBox.information(self, "Success", f"Transactions saved to {file_name}")

            except pymysql.MySQLError as e:
                QMessageBox.critical(self, "Database Error", f"Error fetching transactions: {e}")
            finally:
                if connection and connection.open:
                    cursor.close()
                    connection.close()            

class DrinkMenuWindow(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Drink Menu")
        self.setGeometry(150, 150, 700, 500)
        self.setMinimumSize(600, 400)
        self.init_ui()

    def init_ui(self):
        layout = QVBoxLayout()
        layout.setSpacing(15)
        layout.setContentsMargins(20, 20, 20, 20)

        scroll_area = QScrollArea()
        scroll_area.setWidgetResizable(True)
        # Drink menu table
        self.drink_menu_table = QTableWidget(0, 3)
        self.drink_menu_table.setHorizontalHeaderLabels(["Drink Type", "Variant", "Price (Rp)"])
        self.drink_menu_table.horizontalHeader().setStretchLastSection(True)
        self.drink_menu_table.verticalHeader().setVisible(False)
        self.drink_menu_table.setAlternatingRowColors(True)

        scroll_area.setWidget(self.drink_menu_table)
        layout.addWidget(scroll_area)

        # Button layout
        btn_layout = QHBoxLayout()
        btn_layout.setSpacing(10)

        self.add_btn = QPushButton("Add Drink")
        self.add_btn.clicked.connect(self.add_drink)
        btn_layout.addWidget(self.add_btn)

        self.delete_btn = QPushButton("Delete Drink")
        self.delete_btn.clicked.connect(self.delete_drink)
        btn_layout.addWidget(self.delete_btn)

        layout.addLayout(btn_layout)
        self.setLayout(layout)

        self.load_drinks()

    def load_drinks(self):
        try:
            connection = pymysql.connect(**DB_CONFIG)
            cursor = connection.cursor()
            cursor.execute("SELECT drink_type, variant, price FROM drinks")
            records = cursor.fetchall()

            self.drink_menu_table.setRowCount(0)
            for row_data in records:
                row_count = self.drink_menu_table.rowCount()
                self.drink_menu_table.insertRow(row_count)
                self.drink_menu_table.setItem(row_count, 0, QTableWidgetItem(row_data[0]))
                self.drink_menu_table.setItem(row_count, 1, QTableWidgetItem(row_data[1]))
                self.drink_menu_table.setItem(row_count, 2, QTableWidgetItem(f"Rp {float(row_data[2]):,.0f}"))

        except pymysql.MySQLError as e:
            QMessageBox.critical(self, "Database Error", f"Error loading drinks: {e}")
        finally:
            if connection and connection.open:
                cursor.close()
                connection.close()

    def add_drink(self):
        drink_type, ok1 = QInputDialog.getText(self, "Add Drink", "Enter Drink Type:")
        if not ok1 or not drink_type.strip():
            return

        variant, ok2 = QInputDialog.getText(self, "Add Drink", "Enter Variant:")
        if not ok2 or not variant.strip():
            return

        price, ok3 = QInputDialog.getDouble(self, "Add Drink", "Enter Price:", min=0)
        if ok3:
            try:
                connection = pymysql.connect(**DB_CONFIG)
                cursor = connection.cursor()
                
                cursor.execute(
                    "INSERT INTO drinks (drink_type, variant, price) VALUES (%s, %s, %s)",
                    (drink_type, variant, price)
                )
                connection.commit()
                self.load_drinks()
                
            except pymysql.MySQLError as e:
                QMessageBox.critical(self, "Database Error", f"Error adding drink: {e}")
            finally:
                if connection and connection.open:
                    cursor.close()
                    connection.close()

    def delete_drink(self):
        selected_row = self.drink_menu_table.currentRow()
        if selected_row < 0:
            QMessageBox.warning(self, "Delete Drink", "Please select a drink to delete.")
            return

        drink_type = self.drink_menu_table.item(selected_row, 0).text()
        variant = self.drink_menu_table.item(selected_row, 1).text()

        confirmation = QMessageBox.question(
            self,
            "Confirm Deletion",
            f"Are you sure you want to delete the drink '{drink_type} - {variant}'?",
            QMessageBox.Yes | QMessageBox.No
        )

        if confirmation == QMessageBox.Yes:
            try:
                connection = pymysql.connect(**DB_CONFIG)
                cursor = connection.cursor()
                
                cursor.execute(
                    "DELETE FROM drinks WHERE drink_type = %s AND variant = %s",
                    (drink_type, variant)
                )
                connection.commit()
                self.drink_menu_table.removeRow(selected_row)
                
            except pymysql.MySQLError as e:
                QMessageBox.critical(self, "Database Error", f"Error deleting drink: {e}")
            finally:
                if connection and connection.open:
                    cursor.close()
                    connection.close()


if __name__ == "__main__":
    try:
        print("Starting application...")
        
        # Initialize database
        try:
            create_database()
        except Exception as e:
            print(f"Failed to initialize database: {e}")
            sys.exit(1)
            
        print("Database initialization completed")
        
        # Start GUI application
        try:
            print("Starting GUI...")
            app = QApplication(sys.argv)
            main_window = MainMenu()
            main_window.show()
            print("GUI started successfully")
            sys.exit(app.exec_())
        except Exception as e:
            print(f"Failed to start GUI: {e}")
            sys.exit(1)
            
    except Exception as e:
        print(f"Unexpected error: {e}")
        sys.exit(1)